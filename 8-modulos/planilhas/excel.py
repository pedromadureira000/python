"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
import openpyxl
from random import uniform

"""
pedidos = openpyxl.load_workbook('pedidos.xlsx')  # coleção de abas.
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

print(planilha1['b4'].value)  # acessando o valor diretamente

for campo in planilha1['b']:  # obtendo os valores da coluna
    print(campo.value)

for linha in planilha1[a1:c2]:  # iterando em um range por linhas
    for coluna in linha:
        print(coluna.value) 

for linha in planilha1:  # iterando diretamente na planilha(nao fica muito bom)
    for coluna in linha:
        print(coluna.value)  

planilha1[b3].value = 2200  # alterando dados no objeto que criamos
pedidos.save("nova_planilha.xlsx")  # criando uma planilha com esse objeto

for linha in range(5, 16):  # alterando valores de um range de linha usando "cell".
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido  # o segundo argumento é o numero da coluna
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save('nova_planilha.xlsx')
"""

planilha = openpyxl.Workbook()  # cria um workbook
planilha.create_sheet('Planilha1', 0)  # cria uma planilha na aba 0
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):  # adicionando os dados
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Luiz {linha} {round(uniform(10, 100), 2)}'  # cria nº aliatorio
    planilha2.cell(linha, 2).value = f'Otávio {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Joãozinho {linha} {round(uniform(10, 100), 2)}'

planilha.save('nova_planilha.xlsx')

